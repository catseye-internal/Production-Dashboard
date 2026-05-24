#!/usr/bin/env python3
"""
Production Dashboard — Invoice CSV/XLS Loader (one-time seed)

Reads a PestPac invoice export and writes a curated cache-invoices.json. Two
input formats are supported automatically:

  1. **PestPac quick-export** (tab-delimited despite .xls extension).
     Columns named: "Invoice #", "Sub-Total", "Bill-To Code", "Tech 2", etc.
     Has Excel-protection wrappers: ="06281-1437".

  2. **PestPac Report Writer CSV/Excel export** (comma-delimited .csv, or
     real .xlsx binary). Columns named: "Invoice Number", "SubTotal",
     "BillToID", "Tech 2 Code - Tech 2", "Company / First Name / Last Name",
     etc. No Excel-protection wrappers.

Both formats produce the same output schema. Missing optional columns are
silently skipped — the loader only requires Invoice #, Invoice Date, Branch,
and Invoice Type to produce a usable record.

USAGE
  python3 parse_invoices.py <input.csv|xls|xlsx> <output.json>

  # default — uses the YTD file Joe uploaded:
  python3 parse_invoices.py

DESIGN
  - Auto-detects format: CSV (comma), TSV/.xls (tab), or .xlsx (openpyxl)
  - Alias-aware column matching: every internal field accepts multiple source
    header names so future schema drift is absorbed without breaking the cache
  - Synthesizes `CustomerName` from Company / First Name / Last Name when the
    single "Name" field isn't present (Report Writer split-name pattern)
  - Strips Excel-protection wrappers when present: ="06281-1437" → "06281-1437"
  - Normalizes dates to ISO: 2026/01/01 → 2026-01-01 (matches API order
    WorkDate format)
  - Coerces numeric strings to numbers where useful
  - Drops empty strings to save cache bytes
"""

import csv
import json
import sys
import re
from datetime import datetime
from pathlib import Path

# ── Alias-aware column map ──
# Every internal field key maps to a LIST of possible source header names.
# Order = priority (first match wins). Add new aliases when a new export
# format introduces a different label for an existing field.
FIELD_ALIASES = {
    'InvoiceNumber':      ['Invoice #', 'Invoice Number', 'InvoiceNumber'],
    'InvoiceType':        ['Invoice Type', 'InvoiceType'],
    'OrderNumber':        ['Order #', 'Order Number', 'OrderNumber'],
    'InvoiceDate':        ['Invoice Date', 'InvoiceDate'],
    'WorkDate':           ['Work Date', 'WorkDate'],
    'OrderDate':          ['Order Date', 'OrderDate'],
    'Branch':             ['Branch'],
    'Region':             ['Branch Region', 'Region'],
    'SubRegion':          ['Branch Sub-Region', 'SubRegion'],
    'LocationCode':       ['Location', 'Location Code', 'LocationCode', 'LocID'],
    'BillToCode':         ['Bill-To Code', 'BillToID', 'BillToCode', 'Bill To'],
    'CustomerName':       ['Name', 'Customer', 'Customer Name'],
    'City':               ['City'],
    'State':              ['State'],
    'Zip':                ['Zip code', 'Zip Code', 'Zip', 'ZIP'],
    'Tech':               ['Tech', 'Tech 1', 'Tech 1 Code - Tech', 'Tech 1 Code', 'Tech1'],
    'Tech2':              ['Tech 2', 'Tech 2 Code - Tech 2', 'Tech 2 Code', 'Tech2'],
    'Sales':              ['Sales', 'Tech 3 Code - Sales', 'Tech 3 Code'],
    'EnteredBy':          ['Entered', 'Tech 4 Code - Entered', 'Tech 4 Code', 'EnteredBy'],
    'ServiceClass':       ['Service Class', 'Service Class Code', 'ServiceClass'],
    'ServiceDescription': ['Service Description', 'Service Class Description', 'ServiceDescription'],
    'ServiceCode':        ['Service', 'Service Code', 'ServiceCode'],
    'Route':              ['Route'],
    'SubTotal':           ['Sub-Total', 'SubTotal', 'Sub Total'],
    'Tax':                ['Tax'],
    'Total':              ['Total'],
    'Balance':            ['Balance'],
    'AgingDays':          ['Aging Days', 'AgingDays'],
    'NetDays':            ['Net', 'Terms Net Days', 'NetDays', 'Net Days'],
    'SaleValue':          ['Sale Value', 'SaleValue'],
    'ProductionValue':    ['Production Value', 'ProductionValue'],
    'TaxableAmount':      ['Taxable Amount', 'TaxableAmount'],
    'TaxRate':            ['Tax Rate', 'TaxRate'],
    'Source':             ['Source'],
    'Origin':             ['Origin'],
    'PostedBy':           ['Posted By', 'PostedBy', 'Add User'],  # Add User is RW's closest analog
    'OrderType':          ['Order Type', 'OrderType'],
    'Duration':           ['Duration'],
    'Frequency':          ['Frequency'],
    'Schedule':           ['Schedule'],
    'SetupType':          ['Setup Type', 'SetupType'],
    'GLCode':             ['GL Code', 'GLCode'],
    'InvoiceBatch':       ['Invoice Batch', 'InvoiceBatch'],
    'ConsolidatedNum':    ['Consolidated Num', 'Consolidated Invoice Number', 'ConsolidatedNum'],
    'ConsolidatedDate':   ['Consolidated Date', 'Consolidated Invoice Date', 'ConsolidatedDate'],
    'CreditReason':       ['Credit Reason', 'Credit Reason Code', 'CreditReason'],
    'TaxCode':            ['Tax Code', 'TaxCode'],
    'AccountType':        ['Type', 'AccountType'],   # R/C — Residential/Commercial
}

# Backward-compat alias map for callers that imported the old FIELD_MAP.
# Keys: source header name → output key. First-listed alias for each field wins.
FIELD_MAP = {alias: out_key for out_key, aliases in FIELD_ALIASES.items() for alias in aliases}

# Headers used for the auto-name-synthesis path (Report Writer split-name pattern)
SPLIT_NAME_HEADERS = ('Company', 'First Name', 'Last Name')

NUMERIC_FIELDS = {
    'SubTotal', 'Tax', 'Total', 'Balance', 'AgingDays', 'NetDays',
    'SaleValue', 'ProductionValue', 'TaxableAmount', 'TaxRate', 'Duration',
}

DATE_FIELDS = {
    'InvoiceDate', 'WorkDate', 'OrderDate', 'ConsolidatedDate', 'AddDate',
}

# Required fields — if a row is missing any of these, the record is dropped
REQUIRED_FIELDS = ('InvoiceNumber', 'InvoiceType', 'Branch', 'InvoiceDate')

# Invoice Type lens (Joe's directive 2026-05-20):
# The cache keeps EVERY type. The dashboard toggles between three lenses at
# render time, so we don't have to re-pull the cache when switching views.
COMPLETED_INVOICE_TYPES = None  # disabled — keep all types in cache

# Excel anti-truncation wrapper: ="06281-1437"  →  06281-1437
EXCEL_WRAP = re.compile(r'^="(.*)"$')

KEY_HEADER_TOKENS = ('Invoice #', 'Invoice Number', 'InvoiceNumber')


# ── Cell utilities ──

def clean(v):
    if v is None:
        return None
    s = str(v).strip()
    if not s:
        return ''
    m = EXCEL_WRAP.match(s)
    if m:
        s = m.group(1)
    return s

def to_number(s):
    if s is None or s == '':
        return None
    try:
        return float(str(s).replace(',', '').replace('$', ''))
    except (ValueError, AttributeError):
        return None

def to_iso_date(s):
    if not s:
        return ''
    s = str(s).strip()
    # Strip time portion if present
    if 'T' in s:
        s = s.split('T', 1)[0]
    if ' ' in s and ':' in s:
        s = s.split(' ', 1)[0]
    for fmt in ('%Y/%m/%d', '%m/%d/%Y', '%Y-%m-%d', '%m/%d/%y'):
        try:
            return datetime.strptime(s, fmt).strftime('%Y-%m-%d')
        except ValueError:
            continue
    return s  # leave as-is if unparseable


# ── Format detection ──

def detect_format(path):
    """Returns one of 'csv', 'tsv', 'xlsx'."""
    p = Path(path)
    ext = p.suffix.lower()
    if ext == '.xlsx':
        return 'xlsx'
    # Sniff the first non-empty line — tab or comma?
    with open(path, 'r', encoding='utf-8', errors='replace') as f:
        for line in f:
            line = line.rstrip('\r\n')
            if not line:
                continue
            tabs = line.count('\t')
            # Counting commas only OUTSIDE quoted fields is hard; rough heuristic is fine.
            commas = line.count(',')
            return 'tsv' if tabs >= max(commas, 1) else 'csv'
    return 'csv'

def find_header_row_text(path, delimiter, max_scan=50):
    """Text formats (CSV/TSV) — find the row that looks like the column header."""
    best = (0, -1)
    with open(path, 'r', encoding='utf-8', errors='replace', newline='') as f:
        reader = csv.reader(f, delimiter=delimiter)
        for i, cells in enumerate(reader):
            if i >= max_scan:
                break
            cells = [c.strip() for c in cells]
            if any(k in cells for k in KEY_HEADER_TOKENS) and len(cells) > best[0]:
                best = (len(cells), i)
    if best[1] < 0:
        raise RuntimeError(f"Could not locate header row in {path}")
    return best[1]

# Backward-compat — the old name some callers (merge_invoices.py) used
find_header_row = find_header_row_text


# ── Row iteration adaptors ──

def iter_rows_text(path, delimiter):
    """Yields (header_list, data_row_iter)."""
    header_row = find_header_row_text(path, delimiter)
    f = open(path, 'r', encoding='utf-8', errors='replace', newline='')
    for _ in range(header_row):
        f.readline()
    reader = csv.reader(f, delimiter=delimiter)
    headers = [h.strip() for h in next(reader)]
    return headers, reader, f

def iter_rows_xlsx(path):
    """Real binary xlsx via openpyxl."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise RuntimeError("openpyxl not installed — run `pip3 install openpyxl --break-system-packages`")
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    # Find header row
    header_idx = -1
    for i, r in enumerate(rows[:50]):
        cells = [str(c).strip() if c is not None else '' for c in r]
        if any(k in cells for k in KEY_HEADER_TOKENS):
            header_idx = i
            break
    if header_idx < 0:
        raise RuntimeError(f"Could not locate header row in {path}")
    headers = [str(c).strip() if c is not None else '' for c in rows[header_idx]]
    # Yield remaining rows as lists of strings
    def gen():
        for r in rows[header_idx + 1:]:
            yield [('' if c is None else str(c)) for c in r]
    return headers, gen(), None


# ── Column resolution ──

def resolve_column_indices(headers):
    """
    Returns:
      keep: list of (col_idx, output_field_key)
      split_name_idx: dict {'Company': idx, 'First Name': idx, 'Last Name': idx}
        — only set if the report uses split-name and DOESN'T have a single
        Name/Customer field. Used to synthesize CustomerName.
    """
    # Build header_index map (lowercase normalized for slight forgiveness)
    norm = lambda s: s.strip().lower()
    idx_by_name = {}
    for i, h in enumerate(headers):
        if h:
            idx_by_name.setdefault(h, i)
            idx_by_name.setdefault(norm(h), i)

    keep = []
    matched_keys = set()
    for out_key, aliases in FIELD_ALIASES.items():
        for a in aliases:
            i = idx_by_name.get(a)
            if i is None:
                i = idx_by_name.get(norm(a))
            if i is not None:
                keep.append((i, out_key))
                matched_keys.add(out_key)
                break

    # Synthesize CustomerName from split fields if the single Name field is absent
    split_name_idx = {}
    if 'CustomerName' not in matched_keys:
        for h in SPLIT_NAME_HEADERS:
            i = idx_by_name.get(h)
            if i is not None:
                split_name_idx[h] = i
    return keep, split_name_idx


def synth_customer_name(row, split_name_idx):
    """Concat split name fields. Prefers Company (commercial), falls back to First + Last."""
    company = clean(row[split_name_idx['Company']]) if 'Company' in split_name_idx and len(row) > split_name_idx['Company'] else ''
    if company:
        return company
    first = clean(row[split_name_idx['First Name']]) if 'First Name' in split_name_idx and len(row) > split_name_idx['First Name'] else ''
    last  = clean(row[split_name_idx['Last Name']])  if 'Last Name'  in split_name_idx and len(row) > split_name_idx['Last Name']  else ''
    parts = [p for p in (first, last) if p]
    return ' '.join(parts)


# ── Main parse ──

def parse(input_path, output_path):
    fmt = detect_format(input_path)
    print(f"  Format detected: {fmt}")

    if fmt == 'xlsx':
        headers, reader, fh = iter_rows_xlsx(input_path)
    else:
        delim = '\t' if fmt == 'tsv' else ','
        headers, reader, fh = iter_rows_text(input_path, delim)

    keep, split_name_idx = resolve_column_indices(headers)
    matched_fields = sorted({out_key for _, out_key in keep})
    missing_fields = [k for k in FIELD_ALIASES.keys() if k not in matched_fields]

    print(f"  Total source columns: {len(headers)}  →  curated: {len(keep)}")
    print(f"  Matched fields: {len(matched_fields)} / {len(FIELD_ALIASES)}")
    if missing_fields:
        print(f"  Optional fields not present in this export (skipped): {missing_fields}")
    if split_name_idx:
        print(f"  Customer name will be synthesized from split fields: {list(split_name_idx.keys())}")

    # Find InvoiceType column index for early filtering if needed
    it_alias_indices = []
    for i, h in enumerate(headers):
        if h in FIELD_ALIASES['InvoiceType']:
            it_alias_indices.append(i)
    inv_type_idx = it_alias_indices[0] if it_alias_indices else -1

    records = []
    skipped_short = 0
    skipped_missing_required = 0
    excluded_type = 0
    max_keep_idx = max((i for i, _ in keep), default=-1)

    for row in reader:
        if not row:
            continue
        # Skip empty rows
        if all((c is None or str(c).strip() == '') for c in row):
            continue
        if len(row) < max_keep_idx + 1:
            skipped_short += 1
            continue
        # Skip report-title rows that sometimes precede the real header but
        # also occur as orphan summary rows (e.g. "Invoices_Rolling60")
        if str(row[0] or '').strip().startswith(('Invoice Date:', 'Total:', 'Branch:', 'Subtotal')):
            continue

        # Early type filter (disabled by default — dashboard handles lens at render)
        if COMPLETED_INVOICE_TYPES is not None and inv_type_idx >= 0:
            t = clean(row[inv_type_idx])
            if t and t not in COMPLETED_INVOICE_TYPES:
                excluded_type += 1
                continue

        rec = {}
        for i, out_key in keep:
            val = clean(row[i])
            if out_key in NUMERIC_FIELDS:
                n = to_number(val)
                if n is None or (n == 0 and out_key not in ('Total',)):
                    continue
                rec[out_key] = n
            elif out_key in DATE_FIELDS:
                d = to_iso_date(val)
                if d:
                    rec[out_key] = d
            else:
                if val:
                    rec[out_key] = val

        # Synthesize CustomerName from split fields if needed
        if split_name_idx and 'CustomerName' not in rec:
            cn = synth_customer_name(row, split_name_idx)
            if cn:
                rec['CustomerName'] = cn

        # Drop rows that don't have the bare-minimum identifying fields
        if not all(k in rec for k in REQUIRED_FIELDS):
            skipped_missing_required += 1
            continue

        records.append(rec)

    if fh is not None:
        fh.close()

    print(f"  Parsed: {len(records):,} records")
    print(f"  Skipped (short rows): {skipped_short}")
    print(f"  Skipped (missing required field): {skipped_missing_required}")
    if excluded_type:
        print(f"  Excluded by type filter: {excluded_type:,}")

    # Stats
    branches = sorted({r.get('Branch','') for r in records if r.get('Branch')})
    types    = sorted({r.get('InvoiceType','') for r in records if r.get('InvoiceType')})
    by_branch = {b: sum(1 for r in records if r.get('Branch') == b) for b in branches}
    total_balance = sum(r.get('Balance') or 0 for r in records)
    total_revenue = sum(r.get('Total') or 0 for r in records)

    print(f"  Branches: {branches}")
    print(f"  Invoice types: {types}")
    print(f"  Records by branch: {by_branch}")
    print(f"  Revenue (sum of Total): ${total_revenue:,.2f}")
    print(f"  Outstanding balance: ${total_balance:,.2f}")

    payload = {
        'updated':     datetime.utcnow().isoformat() + 'Z',
        'source':      'seed-from-pestpac-report',
        'sourceFile':  Path(input_path).name,
        'sourceFormat': fmt,
        'recordCount': len(records),
        'invoices':    records,
    }

    out_path = Path(output_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with open(out_path, 'w', encoding='utf-8') as out:
        json.dump(payload, out, separators=(',', ':'))
    size_mb = out_path.stat().st_size / 1024 / 1024
    print(f"  Wrote {out_path}  ({size_mb:.2f} MB)")
    return len(records)


# Backward-compat exports for merge_invoices.py
def parse_report(input_path):
    """Return a list of curated invoice dicts WITHOUT writing to disk.
    Used by merge_invoices.py for upserting into the live cache."""
    fmt = detect_format(input_path)
    print(f"  Format detected: {fmt}")

    if fmt == 'xlsx':
        headers, reader, fh = iter_rows_xlsx(input_path)
    else:
        delim = '\t' if fmt == 'tsv' else ','
        headers, reader, fh = iter_rows_text(input_path, delim)

    keep, split_name_idx = resolve_column_indices(headers)
    matched_fields = sorted({out_key for _, out_key in keep})
    missing_fields = [k for k in FIELD_ALIASES.keys() if k not in matched_fields]
    print(f"  Total source columns: {len(headers)}  →  curated: {len(keep)}")
    print(f"  Matched fields: {len(matched_fields)} / {len(FIELD_ALIASES)}")
    if missing_fields:
        print(f"  Optional fields not present (skipped): {missing_fields}")
    if split_name_idx:
        print(f"  CustomerName will be synthesized from: {list(split_name_idx.keys())}")

    records = []
    skipped_short = 0
    skipped_missing_required = 0
    max_keep_idx = max((i for i, _ in keep), default=-1)

    for row in reader:
        if not row:
            continue
        if all((c is None or str(c).strip() == '') for c in row):
            continue
        if len(row) < max_keep_idx + 1:
            skipped_short += 1
            continue
        if str(row[0] or '').strip().startswith(('Invoice Date:', 'Total:', 'Branch:', 'Subtotal')):
            continue
        rec = {}
        for i, out_key in keep:
            val = clean(row[i])
            if out_key in NUMERIC_FIELDS:
                n = to_number(val)
                if n is None or (n == 0 and out_key not in ('Total',)):
                    continue
                rec[out_key] = n
            elif out_key in DATE_FIELDS:
                d = to_iso_date(val)
                if d:
                    rec[out_key] = d
            else:
                if val:
                    rec[out_key] = val
        if split_name_idx and 'CustomerName' not in rec:
            cn = synth_customer_name(row, split_name_idx)
            if cn:
                rec['CustomerName'] = cn
        if not all(k in rec for k in REQUIRED_FIELDS):
            skipped_missing_required += 1
            continue
        records.append(rec)
    if fh is not None:
        fh.close()
    print(f"  Parsed: {len(records):,} records  (skipped {skipped_short} short, {skipped_missing_required} missing-required)")
    return records


if __name__ == '__main__':
    if len(sys.argv) >= 3:
        in_path, out_path = sys.argv[1], sys.argv[2]
    else:
        # Defaults — workspace paths
        in_path = '/Users/joedingwall/Desktop/Claude Assets/PRODUCTION DASHBOARD/Invoices 1.1.2026-5.19.2026.xls'
        out_path = '/Users/joedingwall/Desktop/Claude Assets/PRODUCTION DASHBOARD/PRODUCTION DASHBOARD/cache-invoices.json'

    print(f"🧾 Invoice loader")
    print(f"   in:  {in_path}")
    print(f"   out: {out_path}")
    n = parse(in_path, out_path)
    print(f"✅ Done — {n:,} invoices loaded")
